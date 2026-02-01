#!/usr/bin/env python3
"""
Merge two Excel accounting ledger files into one combined file.
Preserves original layout and adds a source column to identify origin.
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows


def read_excel_with_structure(filepath, source_name):
    """Read Excel file and preserve structure, adding source identifier."""
    df = pd.read_excel(filepath, header=None)
    return df, source_name


def merge_accounting_files(file1_path, file2_path, output_path):
    """
    Merge two accounting Excel files into one.
    Adds a 'Source' column to identify which file each record came from.
    """
    # Read both files
    df1 = pd.read_excel(file1_path, header=None)
    df2 = pd.read_excel(file2_path, header=None)

    # Extract company names from file paths
    source1 = file1_path.replace('.xlsx', '').split('/')[-1]
    source2 = file2_path.replace('.xlsx', '').split('/')[-1]

    print(f"Reading: {source1} ({len(df1)} rows)")
    print(f"Reading: {source2} ({len(df2)} rows)")

    # Add source column to both dataframes
    df1['Source'] = source1
    df2['Source'] = source2

    # Get header rows (first 3 rows contain account info and column headers)
    # Row 0: Account code + description
    # Row 1: Company code + name
    # Row 2: Empty row
    # Row 3: Column headers (Налог, Дата, etc.)
    # Row 4+: Actual data

    # Get data rows (from row 4 onwards, skipping the header row at index 3)
    data1 = df1.iloc[4:].copy()
    data2 = df2.iloc[4:].copy()

    # Reset index for data
    data1 = data1.reset_index(drop=True)
    data2 = data2.reset_index(drop=True)

    # Combine data rows
    combined_data = pd.concat([data1, data2], ignore_index=True)

    # Sort by date (column 1 contains dates)
    combined_data[1] = pd.to_datetime(combined_data[1], errors='coerce')
    combined_data = combined_data.sort_values(by=1, na_position='first').reset_index(drop=True)

    # Create output workbook with formatting
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Merged Data"

    # Define styles
    header_font = Font(bold=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    source1_fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")  # Light blue
    source2_fill = PatternFill(start_color="FFF3E6", end_color="FFF3E6", fill_type="solid")  # Light orange

    # Write merged header
    ws.cell(row=1, column=1, value="2200")
    ws.cell(row=1, column=2, value="Обврски спрема добавувачи врз основа на набавка на добра (производи) и услуги во земјата")
    ws.cell(row=1, column=11, value="Извор")

    # Write company info
    ws.cell(row=2, column=1, value=f"{source1} (6) + {source2} (40)")
    ws.cell(row=2, column=2, value="ТП БИЛАНС ЕЛИТ - COMBINED")

    # Write column headers
    headers = ["Налог", "Дата", "Вал.", "м.ддв", "Опис", "Затворање", "Забелешка", "Долгува", "Побарува", "Един", "Извор"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    # Write data rows
    row_num = 5
    for idx, row in combined_data.iterrows():
        source = row['Source']
        fill = source1_fill if source == source1 else source2_fill

        for col in range(10):  # Original columns
            value = row.iloc[col]
            # Format date
            if col == 1 and pd.notna(value):
                if isinstance(value, pd.Timestamp):
                    value = value.strftime('%Y-%m-%d')
            cell = ws.cell(row=row_num, column=col+1, value=value)
            cell.border = thin_border
            cell.fill = fill

        # Source column
        cell = ws.cell(row=row_num, column=11, value=source)
        cell.border = thin_border
        cell.fill = fill

        row_num += 1

    # Add summary section
    row_num += 2
    ws.cell(row=row_num, column=1, value="SUMMARY")
    ws.cell(row=row_num, column=1).font = header_font

    row_num += 1
    ws.cell(row=row_num, column=1, value=f"Total records from {source1}:")
    ws.cell(row=row_num, column=2, value=len(data1))

    row_num += 1
    ws.cell(row=row_num, column=1, value=f"Total records from {source2}:")
    ws.cell(row=row_num, column=2, value=len(data2))

    row_num += 1
    ws.cell(row=row_num, column=1, value="Combined total:")
    ws.cell(row=row_num, column=2, value=len(combined_data))

    # Find overlapping Налог codes
    nalog1 = set(data1.iloc[:, 0].dropna().unique())
    nalog2 = set(data2.iloc[:, 0].dropna().unique())
    overlapping = nalog1.intersection(nalog2)

    row_num += 2
    ws.cell(row=row_num, column=1, value="Overlapping Налог codes:")
    ws.cell(row=row_num, column=1).font = header_font
    row_num += 1
    for code in sorted(overlapping):
        ws.cell(row=row_num, column=1, value=code)
        row_num += 1

    # Adjust column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 8
    ws.column_dimensions['K'].width = 15

    # Save workbook
    wb.save(output_path)
    print(f"\nMerged file saved to: {output_path}")
    print(f"Total combined records: {len(combined_data)}")
    print(f"Overlapping Налог codes: {len(overlapping)}")
    if overlapping:
        print(f"Codes: {sorted(overlapping)}")

    # Prepare data for HTML
    csv_columns = ["Налог", "Дата", "Вал", "м_ддв", "Опис", "Затворање", "Забелешка", "Долгува", "Побарува", "Един", "Извор"]
    html_data = combined_data.copy()
    html_data.columns = csv_columns
    html_data['Дата'] = pd.to_datetime(html_data['Дата'], errors='coerce').dt.strftime('%Y-%m-%d')

    # Generate HTML file in prototype folder
    generate_html(html_data, source1, source2, overlapping)

    return combined_data, overlapping


def generate_html(data, source1, source2, overlapping):
    """Generate interactive HTML file for accountants."""

    # Convert data to JSON for JavaScript
    data_json = data.to_json(orient='records', force_ascii=False)

    html_content = f'''<!DOCTYPE html>
<html lang="mk">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Сметководствена книга - Споени податоци</title>
    <style>
        * {{
            box-sizing: border-box;
            font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif;
        }}
        body {{
            margin: 0;
            padding: 20px;
            background: #f5f5f5;
        }}
        .header {{
            background: #2c3e50;
            color: white;
            padding: 20px;
            border-radius: 8px;
            margin-bottom: 20px;
        }}
        .header h1 {{
            margin: 0 0 10px 0;
            font-size: 24px;
        }}
        .header p {{
            margin: 0;
            opacity: 0.8;
        }}
        .controls {{
            background: white;
            padding: 20px;
            border-radius: 8px;
            margin-bottom: 20px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }}
        .filters {{
            display: flex;
            flex-wrap: wrap;
            gap: 15px;
            margin-bottom: 15px;
        }}
        .filter-group {{
            display: flex;
            flex-direction: column;
            gap: 5px;
        }}
        .filter-group label {{
            font-size: 12px;
            font-weight: 600;
            color: #666;
        }}
        .filter-group input, .filter-group select {{
            padding: 8px 12px;
            border: 1px solid #ddd;
            border-radius: 4px;
            font-size: 14px;
        }}
        .summary-cards {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 15px;
            margin-bottom: 20px;
        }}
        .card {{
            background: white;
            padding: 20px;
            border-radius: 8px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }}
        .card.blue {{ border-left: 4px solid #3498db; }}
        .card.orange {{ border-left: 4px solid #e67e22; }}
        .card.green {{ border-left: 4px solid #27ae60; }}
        .card.red {{ border-left: 4px solid #e74c3c; }}
        .card h3 {{
            margin: 0 0 10px 0;
            font-size: 14px;
            color: #666;
        }}
        .card .value {{
            font-size: 28px;
            font-weight: bold;
            color: #2c3e50;
        }}
        .card .subtitle {{
            font-size: 12px;
            color: #999;
            margin-top: 5px;
        }}
        .table-container {{
            background: white;
            border-radius: 8px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
            overflow: hidden;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
        }}
        th {{
            background: #34495e;
            color: white;
            padding: 12px 8px;
            text-align: left;
            font-size: 13px;
            cursor: pointer;
            user-select: none;
            position: sticky;
            top: 0;
        }}
        th:hover {{
            background: #2c3e50;
        }}
        th .sort-icon {{
            margin-left: 5px;
            opacity: 0.5;
        }}
        th.sorted .sort-icon {{
            opacity: 1;
        }}
        td {{
            padding: 10px 8px;
            border-bottom: 1px solid #eee;
            font-size: 13px;
        }}
        tr:hover {{
            background: #f8f9fa;
        }}
        tr.source-hami {{
            background: #e8f4fc;
        }}
        tr.source-zubeks {{
            background: #fef5e7;
        }}
        tr.source-hami:hover {{
            background: #d4e9f7;
        }}
        tr.source-zubeks:hover {{
            background: #fdebd0;
        }}
        .number {{
            text-align: right;
            font-family: monospace;
        }}
        .totals-row {{
            background: #2c3e50 !important;
            color: white;
            font-weight: bold;
        }}
        .totals-row td {{
            border-bottom: none;
        }}
        .badge {{
            display: inline-block;
            padding: 3px 8px;
            border-radius: 12px;
            font-size: 11px;
            font-weight: 600;
        }}
        .badge-hami {{
            background: #3498db;
            color: white;
        }}
        .badge-zubeks {{
            background: #e67e22;
            color: white;
        }}
        .overlap-indicator {{
            color: #e74c3c;
            font-weight: bold;
        }}
        .no-data {{
            text-align: center;
            padding: 40px;
            color: #999;
        }}
        .export-btn {{
            background: #27ae60;
            color: white;
            border: none;
            padding: 10px 20px;
            border-radius: 4px;
            cursor: pointer;
            font-size: 14px;
        }}
        .export-btn:hover {{
            background: #219a52;
        }}
        .reset-btn {{
            background: #95a5a6;
            color: white;
            border: none;
            padding: 10px 20px;
            border-radius: 4px;
            cursor: pointer;
            font-size: 14px;
            margin-left: 10px;
        }}
        .reset-btn:hover {{
            background: #7f8c8d;
        }}
        .stats-section {{
            background: white;
            padding: 20px;
            border-radius: 8px;
            margin-bottom: 20px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }}
        .stats-section h2 {{
            margin: 0 0 15px 0;
            font-size: 18px;
            color: #2c3e50;
            border-bottom: 2px solid #3498db;
            padding-bottom: 10px;
        }}
        .stats-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(220px, 1fr));
            gap: 15px;
        }}
        .stat-box {{
            background: #f8f9fa;
            padding: 15px;
            border-radius: 6px;
            border-left: 3px solid #bdc3c7;
        }}
        .stat-box.highlight {{
            border-left-color: #9b59b6;
            background: #f5eef8;
        }}
        .stat-box.source1-box {{
            border-left-color: #3498db;
            background: #eaf2f8;
        }}
        .stat-box.source2-box {{
            border-left-color: #e67e22;
            background: #fef5e7;
        }}
        .stat-label {{
            font-size: 12px;
            color: #666;
            font-weight: 600;
            margin-bottom: 5px;
        }}
        .stat-value {{
            font-size: 24px;
            font-weight: bold;
            color: #2c3e50;
            font-family: monospace;
        }}
        .stat-value-small {{
            font-size: 16px;
            font-weight: bold;
            color: #2c3e50;
        }}
        .stat-detail {{
            font-size: 11px;
            color: #888;
            margin-top: 5px;
        }}
        .stat-value.positive {{
            color: #27ae60;
        }}
        .stat-value.negative {{
            color: #e74c3c;
        }}
    </style>
</head>
<body>
    <div class="header">
        <h1>Сметководствена книга - Споени податоци</h1>
        <p>Конто 2200: Обврски спрема добавувачи | {source1} + {source2}</p>
    </div>

    <div class="summary-cards">
        <div class="card blue">
            <h3>Записи од {source1}</h3>
            <div class="value" id="count-source1">0</div>
            <div class="subtitle">Побарува: <span id="sum-source1">0</span> | Долгува: <span id="sum-dolgува-source1">0</span></div>
        </div>
        <div class="card orange">
            <h3>Записи од {source2}</h3>
            <div class="value" id="count-source2">0</div>
            <div class="subtitle">Побарува: <span id="sum-source2">0</span> | Долгува: <span id="sum-dolgува-source2">0</span></div>
        </div>
        <div class="card green">
            <h3>Вкупно прикажани</h3>
            <div class="value" id="count-total">0</div>
            <div class="subtitle">Долгува: <span id="sum-dolgува">0</span> | Побарува: <span id="sum-pobarува">0</span></div>
        </div>
        <div class="card red">
            <h3>Преклопени кодови</h3>
            <div class="value">{len(overlapping)}</div>
            <div class="subtitle">Записи кои се појавуваат во двете датотеки</div>
        </div>
    </div>

    <div class="stats-section">
        <h2>Финансиски преглед</h2>
        <div class="stats-grid">
            <div class="stat-box">
                <div class="stat-label">Вкупно фактури (Побарува)</div>
                <div class="stat-value" id="total-invoices">0</div>
                <div class="stat-detail">Број на фактури: <span id="invoice-count">0</span></div>
            </div>
            <div class="stat-box">
                <div class="stat-label">Вкупно плаќања (Долгува)</div>
                <div class="stat-value" id="total-payments">0</div>
                <div class="stat-detail">Број на плаќања: <span id="payment-count">0</span></div>
            </div>
            <div class="stat-box highlight">
                <div class="stat-label">Салдо (Побарува - Долгува)</div>
                <div class="stat-value" id="balance">0</div>
                <div class="stat-detail" id="balance-status">Неподмирено</div>
            </div>
            <div class="stat-box">
                <div class="stat-label">Период на податоци</div>
                <div class="stat-value-small" id="date-range">-</div>
                <div class="stat-detail">Број на месеци: <span id="month-count">0</span></div>
            </div>
        </div>
        <div class="stats-grid" style="margin-top: 15px;">
            <div class="stat-box source1-box">
                <div class="stat-label">Салдо {source1}</div>
                <div class="stat-value" id="balance-source1">0</div>
                <div class="stat-detail">Фактури: <span id="inv-source1">0</span> | Плаќања: <span id="pay-source1">0</span></div>
            </div>
            <div class="stat-box source2-box">
                <div class="stat-label">Салдо {source2}</div>
                <div class="stat-value" id="balance-source2">0</div>
                <div class="stat-detail">Фактури: <span id="inv-source2">0</span> | Плаќања: <span id="pay-source2">0</span></div>
            </div>
            <div class="stat-box">
                <div class="stat-label">Просечна фактура</div>
                <div class="stat-value" id="avg-invoice">0</div>
                <div class="stat-detail">Мин: <span id="min-invoice">0</span> | Макс: <span id="max-invoice">0</span></div>
            </div>
            <div class="stat-box">
                <div class="stat-label">Просечно плаќање</div>
                <div class="stat-value" id="avg-payment">0</div>
                <div class="stat-detail">Мин: <span id="min-payment">0</span> | Макс: <span id="max-payment">0</span></div>
            </div>
        </div>
    </div>

    <div class="controls">
        <div class="filters">
            <div class="filter-group">
                <label>Пребарај (Налог, Опис)</label>
                <input type="text" id="searchInput" placeholder="Внесете текст...">
            </div>
            <div class="filter-group">
                <label>Извор</label>
                <select id="sourceFilter">
                    <option value="">Сите извори</option>
                    <option value="{source1}">{source1}</option>
                    <option value="{source2}">{source2}</option>
                </select>
            </div>
            <div class="filter-group">
                <label>Датум од</label>
                <input type="date" id="dateFrom">
            </div>
            <div class="filter-group">
                <label>Датум до</label>
                <input type="date" id="dateTo">
            </div>
            <div class="filter-group">
                <label>Месец (м.ддв)</label>
                <select id="monthFilter">
                    <option value="">Сите месеци</option>
                </select>
            </div>
            <div class="filter-group">
                <label>Прикажи преклопени</label>
                <select id="overlapFilter">
                    <option value="">Сите записи</option>
                    <option value="yes">Само преклопени</option>
                    <option value="no">Само непреклопени</option>
                </select>
            </div>
        </div>
        <button class="export-btn" onclick="exportFiltered()">Извези филтрирани во CSV</button>
        <button class="reset-btn" onclick="resetFilters()">Ресетирај филтри</button>
    </div>

    <div class="table-container">
        <table id="dataTable">
            <thead>
                <tr>
                    <th data-col="Налог">Налог <span class="sort-icon">↕</span></th>
                    <th data-col="Дата">Дата <span class="sort-icon">↕</span></th>
                    <th data-col="м_ддв">м.ддв <span class="sort-icon">↕</span></th>
                    <th data-col="Опис">Опис <span class="sort-icon">↕</span></th>
                    <th data-col="Затворање">Затворање <span class="sort-icon">↕</span></th>
                    <th data-col="Забелешка">Забелешка <span class="sort-icon">↕</span></th>
                    <th data-col="Долгува" class="number">Долгува <span class="sort-icon">↕</span></th>
                    <th data-col="Побарува" class="number">Побарува <span class="sort-icon">↕</span></th>
                    <th data-col="Извор">Извор <span class="sort-icon">↕</span></th>
                </tr>
            </thead>
            <tbody id="tableBody">
            </tbody>
            <tfoot>
                <tr class="totals-row">
                    <td colspan="6"><strong>ВКУПНО (Филтрирано)</strong></td>
                    <td class="number" id="total-dolgува">0</td>
                    <td class="number" id="total-pobarува">0</td>
                    <td></td>
                </tr>
            </tfoot>
        </table>
    </div>

    <script>
        const rawData = {data_json};
        const overlappingCodes = {list(overlapping)};
        const source1 = "{source1}";
        const source2 = "{source2}";

        let filteredData = [...rawData];
        let sortCol = 'Дата';
        let sortAsc = true;

        function formatNumber(num) {{
            if (num === null || num === undefined || num === '' || isNaN(num)) return '';
            return Number(num).toLocaleString('en-US');
        }}

        function parseNumber(val) {{
            if (val === null || val === undefined || val === '') return 0;
            return Number(val) || 0;
        }}

        function populateMonthFilter() {{
            const months = [...new Set(rawData.map(r => r['м_ддв']).filter(m => m))].sort((a,b) => a-b);
            const select = document.getElementById('monthFilter');
            months.forEach(m => {{
                const opt = document.createElement('option');
                opt.value = m;
                opt.textContent = m;
                select.appendChild(opt);
            }});
        }}

        function applyFilters() {{
            const search = document.getElementById('searchInput').value.toLowerCase();
            const source = document.getElementById('sourceFilter').value;
            const dateFrom = document.getElementById('dateFrom').value;
            const dateTo = document.getElementById('dateTo').value;
            const month = document.getElementById('monthFilter').value;
            const overlap = document.getElementById('overlapFilter').value;

            filteredData = rawData.filter(row => {{
                // Search filter
                if (search) {{
                    const nalog = (row['Налог'] || '').toLowerCase();
                    const opis = (row['Опис'] || '').toLowerCase();
                    if (!nalog.includes(search) && !opis.includes(search)) return false;
                }}
                // Source filter
                if (source && row['Извор'] !== source) return false;
                // Date filters
                if (dateFrom && row['Дата'] < dateFrom) return false;
                if (dateTo && row['Дата'] > dateTo) return false;
                // Month filter
                if (month && row['м_ддв'] != month) return false;
                // Overlap filter
                if (overlap === 'yes' && !overlappingCodes.includes(row['Налог'])) return false;
                if (overlap === 'no' && overlappingCodes.includes(row['Налог'])) return false;

                return true;
            }});

            sortData();
            renderTable();
            updateSummary();
        }}

        function sortData() {{
            filteredData.sort((a, b) => {{
                let valA = a[sortCol];
                let valB = b[sortCol];

                // Handle numbers
                if (sortCol === 'Долгува' || sortCol === 'Побарува' || sortCol === 'м_ддв') {{
                    valA = parseNumber(valA);
                    valB = parseNumber(valB);
                }}

                if (valA < valB) return sortAsc ? -1 : 1;
                if (valA > valB) return sortAsc ? 1 : -1;
                return 0;
            }});
        }}

        function renderTable() {{
            const tbody = document.getElementById('tableBody');

            if (filteredData.length === 0) {{
                tbody.innerHTML = '<tr><td colspan="9" class="no-data">Нема записи што одговараат на филтрите</td></tr>';
                return;
            }}

            tbody.innerHTML = filteredData.map(row => {{
                const isOverlap = overlappingCodes.includes(row['Налог']);
                const sourceClass = row['Извор'] === source1 ? 'source-hami' : 'source-zubeks';
                const badgeClass = row['Извор'] === source1 ? 'badge-hami' : 'badge-zubeks';

                return `
                    <tr class="${{sourceClass}}">
                        <td>${{row['Налог'] || ''}} ${{isOverlap ? '<span class="overlap-indicator">*</span>' : ''}}</td>
                        <td>${{row['Дата'] || ''}}</td>
                        <td>${{row['м_ддв'] || ''}}</td>
                        <td>${{row['Опис'] || ''}}</td>
                        <td>${{row['Затворање'] || ''}}</td>
                        <td>${{row['Забелешка'] || ''}}</td>
                        <td class="number">${{formatNumber(row['Долгува'])}}</td>
                        <td class="number">${{formatNumber(row['Побарува'])}}</td>
                        <td><span class="badge ${{badgeClass}}">${{row['Извор']}}</span></td>
                    </tr>
                `;
            }}).join('');

            // Update column header styles
            document.querySelectorAll('th').forEach(th => {{
                th.classList.remove('sorted');
                if (th.dataset.col === sortCol) {{
                    th.classList.add('sorted');
                    th.querySelector('.sort-icon').textContent = sortAsc ? '↑' : '↓';
                }} else {{
                    th.querySelector('.sort-icon').textContent = '↕';
                }}
            }});
        }}

        function updateSummary() {{
            const source1Data = filteredData.filter(r => r['Извор'] === source1);
            const source2Data = filteredData.filter(r => r['Извор'] === source2);

            // Basic counts
            document.getElementById('count-source1').textContent = source1Data.length;
            document.getElementById('count-source2').textContent = source2Data.length;
            document.getElementById('count-total').textContent = filteredData.length;

            // Calculate sums
            const sumDolgува = filteredData.reduce((sum, r) => sum + parseNumber(r['Долгува']), 0);
            const sumPobarува = filteredData.reduce((sum, r) => sum + parseNumber(r['Побарува']), 0);
            const sumSource1Pob = source1Data.reduce((sum, r) => sum + parseNumber(r['Побарува']), 0);
            const sumSource2Pob = source2Data.reduce((sum, r) => sum + parseNumber(r['Побарува']), 0);
            const sumSource1Dol = source1Data.reduce((sum, r) => sum + parseNumber(r['Долгува']), 0);
            const sumSource2Dol = source2Data.reduce((sum, r) => sum + parseNumber(r['Долгува']), 0);

            // Update card summaries
            document.getElementById('sum-dolgува').textContent = formatNumber(sumDolgува);
            document.getElementById('sum-pobarува').textContent = formatNumber(sumPobarува);
            document.getElementById('sum-source1').textContent = formatNumber(sumSource1Pob);
            document.getElementById('sum-source2').textContent = formatNumber(sumSource2Pob);
            document.getElementById('sum-dolgува-source1').textContent = formatNumber(sumSource1Dol);
            document.getElementById('sum-dolgува-source2').textContent = formatNumber(sumSource2Dol);
            document.getElementById('total-dolgува').textContent = formatNumber(sumDolgува);
            document.getElementById('total-pobarува').textContent = formatNumber(sumPobarува);

            // Invoices (Побарува > 0) and Payments (Долгува > 0)
            const invoices = filteredData.filter(r => parseNumber(r['Побарува']) > 0);
            const payments = filteredData.filter(r => parseNumber(r['Долгува']) > 0);

            document.getElementById('total-invoices').textContent = formatNumber(sumPobarува);
            document.getElementById('invoice-count').textContent = invoices.length;
            document.getElementById('total-payments').textContent = formatNumber(sumDolgува);
            document.getElementById('payment-count').textContent = payments.length;

            // Balance
            const balance = sumPobarува - sumDolgува;
            const balanceEl = document.getElementById('balance');
            balanceEl.textContent = formatNumber(balance);
            balanceEl.className = 'stat-value ' + (balance > 0 ? 'negative' : balance < 0 ? 'positive' : '');
            document.getElementById('balance-status').textContent = balance > 0 ? 'Неподмирено задолжување' : balance < 0 ? 'Преплата' : 'Подмирено';

            // Source balances
            const balanceSource1 = sumSource1Pob - sumSource1Dol;
            const balanceSource2 = sumSource2Pob - sumSource2Dol;
            const bal1El = document.getElementById('balance-source1');
            const bal2El = document.getElementById('balance-source2');
            bal1El.textContent = formatNumber(balanceSource1);
            bal2El.textContent = formatNumber(balanceSource2);
            bal1El.className = 'stat-value ' + (balanceSource1 > 0 ? 'negative' : balanceSource1 < 0 ? 'positive' : '');
            bal2El.className = 'stat-value ' + (balanceSource2 > 0 ? 'negative' : balanceSource2 < 0 ? 'positive' : '');

            // Invoice/payment counts per source
            document.getElementById('inv-source1').textContent = source1Data.filter(r => parseNumber(r['Побарува']) > 0).length;
            document.getElementById('pay-source1').textContent = source1Data.filter(r => parseNumber(r['Долгува']) > 0).length;
            document.getElementById('inv-source2').textContent = source2Data.filter(r => parseNumber(r['Побарува']) > 0).length;
            document.getElementById('pay-source2').textContent = source2Data.filter(r => parseNumber(r['Долгува']) > 0).length;

            // Average, min, max for invoices
            const invAmounts = invoices.map(r => parseNumber(r['Побарува'])).filter(v => v > 0);
            const payAmounts = payments.map(r => parseNumber(r['Долгува'])).filter(v => v > 0);

            if (invAmounts.length > 0) {{
                document.getElementById('avg-invoice').textContent = formatNumber(Math.round(invAmounts.reduce((a,b) => a+b, 0) / invAmounts.length));
                document.getElementById('min-invoice').textContent = formatNumber(Math.min(...invAmounts));
                document.getElementById('max-invoice').textContent = formatNumber(Math.max(...invAmounts));
            }} else {{
                document.getElementById('avg-invoice').textContent = '0';
                document.getElementById('min-invoice').textContent = '0';
                document.getElementById('max-invoice').textContent = '0';
            }}

            if (payAmounts.length > 0) {{
                document.getElementById('avg-payment').textContent = formatNumber(Math.round(payAmounts.reduce((a,b) => a+b, 0) / payAmounts.length));
                document.getElementById('min-payment').textContent = formatNumber(Math.min(...payAmounts));
                document.getElementById('max-payment').textContent = formatNumber(Math.max(...payAmounts));
            }} else {{
                document.getElementById('avg-payment').textContent = '0';
                document.getElementById('min-payment').textContent = '0';
                document.getElementById('max-payment').textContent = '0';
            }}

            // Date range
            const dates = filteredData.map(r => r['Дата']).filter(d => d).sort();
            if (dates.length > 0) {{
                document.getElementById('date-range').textContent = dates[0] + ' - ' + dates[dates.length - 1];
            }} else {{
                document.getElementById('date-range').textContent = '-';
            }}

            // Month count
            const months = [...new Set(filteredData.map(r => r['м_ддв']).filter(m => m))];
            document.getElementById('month-count').textContent = months.length;
        }}

        function resetFilters() {{
            document.getElementById('searchInput').value = '';
            document.getElementById('sourceFilter').value = '';
            document.getElementById('dateFrom').value = '';
            document.getElementById('dateTo').value = '';
            document.getElementById('monthFilter').value = '';
            document.getElementById('overlapFilter').value = '';
            applyFilters();
        }}

        function exportFiltered() {{
            const headers = ['Налог', 'Дата', 'Вал', 'м_ддв', 'Опис', 'Затворање', 'Забелешка', 'Долгува', 'Побарува', 'Един', 'Извор'];
            const csvContent = [
                headers.join(','),
                ...filteredData.map(row =>
                    headers.map(h => {{
                        let val = row[h] || '';
                        if (typeof val === 'string' && (val.includes(',') || val.includes('"'))) {{
                            val = '"' + val.replace(/"/g, '""') + '"';
                        }}
                        return val;
                    }}).join(',')
                )
            ].join('\\n');

            const blob = new Blob([csvContent], {{ type: 'text/csv;charset=utf-8;' }});
            const link = document.createElement('a');
            link.href = URL.createObjectURL(blob);
            link.download = 'filtered_export.csv';
            link.click();
        }}

        // Event listeners
        document.getElementById('searchInput').addEventListener('input', applyFilters);
        document.getElementById('sourceFilter').addEventListener('change', applyFilters);
        document.getElementById('dateFrom').addEventListener('change', applyFilters);
        document.getElementById('dateTo').addEventListener('change', applyFilters);
        document.getElementById('monthFilter').addEventListener('change', applyFilters);
        document.getElementById('overlapFilter').addEventListener('change', applyFilters);

        document.querySelectorAll('th[data-col]').forEach(th => {{
            th.addEventListener('click', () => {{
                const col = th.dataset.col;
                if (sortCol === col) {{
                    sortAsc = !sortAsc;
                }} else {{
                    sortCol = col;
                    sortAsc = true;
                }}
                sortData();
                renderTable();
            }});
        }});

        // Initialize
        populateMonthFilter();
        applyFilters();
    </script>
</body>
</html>
'''

    with open('accounting_viewer.html', 'w', encoding='utf-8') as f:
        f.write(html_content)
    print(f"HTML viewer saved to: accounting_viewer.html")


if __name__ == "__main__":
    # Define file paths (source files are in source data folder)
    file1 = "source data/Hami stam.xlsx"
    file2 = "source data/Zubeks.xlsx"
    output = "Merged_Accounting.xlsx"

    # Run merge
    combined, overlaps = merge_accounting_files(file1, file2, output)
